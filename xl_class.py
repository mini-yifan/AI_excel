import pandas as pd
import time
from openpyxl import Workbook, load_workbook
from openpyxl.styles import NamedStyle
from datetime import datetime, timedelta
import shutil
import openpyxl
import re
import os
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter, column_index_from_string


# 获取当前日期年月日时分秒
def get_current_date():
    """
    获得当前时间
    """
    current_time = time.localtime()
    year = current_time.tm_year
    month = current_time.tm_mon
    day = current_time.tm_mday
    hour = current_time.tm_hour
    minute = current_time.tm_min
    second = current_time.tm_sec
    return f"{year}_{month:02d}_{day:02d}_{hour:02d}_{minute:02d}_{second:02d}"

# 复制文件
def copy_excel_with_pandas(source_path, destination_path=None, time_c=False):
    """
    复制文件
    :param source_path:
    :param destination_path:
    :param time_c:
    :return:
    """
    if destination_path is None:
        if time_c:
            destination_path = source_path[:-5] + "_copy" + str(get_current_date()) + '.xlsx'
        else:
            destination_path = source_path[:-5] + "_copy.xlsx"
    try:
        # 复制文件并重命名
        shutil.copy(source_path, destination_path)
        print(f"文件已复制: {destination_path}")
        #return destination_path
    except Exception as ex:
        print(f"复制文件时发生错误: {ex}")


# 删除excel文件
def delete_excel_file(file_path):
    """
    删除文件
    :param file_path:
    :return:
    """
    try:
        # 使用os模块删除文件
        os.remove(file_path)
        print(f"文件已删除: {file_path}")
    except Exception as ex:
        print(f"删除文件时发生错误: {ex}")


# 创建一个包含多个工作表的Excel文件， time_list对应初始年月日，结束年月日
def create_excel_with_multiple_sheets(file_path, sheets_num=1, sheets_data=None, time_list=[], copy=False, split_y="_", split_m="_", split_d=""):
    """
    将当前文件修改为一个包含多个工作表的空文件，工作表名称可以是日期，月份，年份，也可以自定义，
    :param file_path:
    :param sheets_num:创建工作表数量，在time_list=[]时执行
    :param sheets_data:
    :param time_list:格式为[2024, 2, 28, 2024, 3, 4]，2024年2月28到2024年3月4日
    :param copy:是否复制文件进行备份
    :param split_y:
    :param split_m:
    :param split_d:
    :return:
    """
    if copy:
        copy_excel_with_pandas(source_path=file_path)
    # 创建一个新的工作簿对象
    wb = Workbook()
    # 移除默认创建的第一个工作表，因为我们将在后面添加自定义的工作表
    wb.remove(wb.active)
    try:
        # 如果sheets_data是一个字典，则将其工作表名称作为键，数据作为值
        if isinstance(sheets_data, dict):
            print("sheets_data is dict")
            for sheet_name, data in sheets_data.items():
                # 为每个工作表名称创建一个新的工作表
                ws = wb.create_sheet(title=sheet_name)
                # 将数据逐行写入工作表
                for row in data:
                    ws.append(row)

        # 如果sheets_data是一个列表，则将其工作表名称作为元素
        elif isinstance(sheets_data, list):
            for sheet_name in sheets_data:
                ws = wb.create_sheet(title=sheet_name)
                print(f"创建工作表: {sheet_name}")

        # 如果sheets_data为None，则创建指定年份的月份工作表
        elif sheets_data is None:
            if time_list==[]:
                sheets_name_list = list(range(sheets_num))
                for sheet_name in sheets_name_list:
                    ws = wb.create_sheet(title=str(sheet_name))
                    print(f"创建工作表: {sheet_name}")

            else:
                start_year, start_month, start_day, end_year, end_month, end_day = time_list

                if start_month == 0:
                    if start_year > end_year:
                        print("开始年份不能晚于结束年份")
                        return

                    current_year = start_year
                    while current_year <= end_year:
                        title_name = str(current_year)+split_y
                        ws = wb.create_sheet(title=title_name)
                        print("创建工作表: ", str(current_year)+split_y)
                        current_year += 1

                elif start_day == 0:
                    # 确保开始日期在结束日期之前或相等
                    if start_year > end_year or (start_year == end_year and start_month > end_month):
                        print("开始月份不能晚于结束月份")
                        return

                    current_year, current_month = start_year, start_month
                    while (current_year < end_year) or (current_year == end_year and current_month <= end_month):
                        title_name = str(current_year) + split_y + str(current_month) + split_m
                        ws = wb.create_sheet(title=title_name)
                        print("创建工作表: ", str(current_year) + split_y + str(current_month) + split_m)
                        # 更新当前月份和年份
                        if current_month == 12:
                            current_month = 1
                            current_year += 1
                        else:
                            current_month += 1

                else:
                    if end_day == 0:
                        end_day = start_day
                    # 将日期元素转换为日期对象
                    start_date = datetime(start_year, start_month, start_day)
                    end_date = datetime(end_year, end_month, end_day)

                    # 确保开始日期在结束日期之前
                    if start_date > end_date:
                        print("开始日期不能晚于结束日期")
                        return

                    current_date = start_date
                    while current_date <= end_date:
                        year, mouth, day = current_date.year, current_date.month, current_date.day
                        title_name = str(year) + split_y + str(mouth) + split_m + str(day) + split_d
                        print(title_name)
                        ws = wb.create_sheet(title=title_name)
                        print("创建工作表: ", str(year) + split_y + str(mouth) + split_m + str(day) + split_d)
                        current_date += timedelta(days=1)

        # 如果sheets_data是一个元组，则将其工作表名称作为元素
        else:
            sheets_data = list(sheets_data)
            for sheet_name in sheets_data:
                ws = wb.create_sheet(title=str(sheet_name))
                print(f"创建工作表: {sheet_name}")

        # 保存工作簿到指定路径
        wb.save(file_path)
        print(f"Excel文件已成功创建，包含多个工作表: {file_path}")

    except Exception as ex:
        print(f"创建Excel文件时发生错误: {ex}")
        return False


# 将源Excel文件src_file_path的第一个工作表复制到目标Excel文件dest_file_path的所有工作表中。
def copy_first_sheet_to_all_sheets(src_file_path, dest_file_path, sheet_i=0):
    """
    将源Excel文件src_file_path的第sheet_i工作表复制到目标Excel文件dest_file_path的所有工作表中，并保持格式不变。
    参数:
    src_file_path (str): 源Excel文件路径。
    dest_file_path (str): 目标Excel文件路径。
    sheet_i (int): 要复制的源工作表的索引，默认为0（第一个工作表）。
    """

    # 判断文件是否存在
    exist_src = os.path.exists(src_file_path)
    exist_dest = os.path.exists(dest_file_path)
    if exist_src and exist_dest:
        pass
    else:
        print("文件不存在")
        return

    # 加载源和目标工作簿
    src_wb = load_workbook(src_file_path)
    dest_wb = load_workbook(dest_file_path)

    # 获取源工作簿的第sheet_i工作表
    src_ws = src_wb.worksheets[sheet_i]

    try:
        for sheet in dest_wb.worksheets:
            # 清空目标工作表的内容
            sheet.delete_rows(1, sheet.max_row)

            # 复制源工作表的内容到目标工作表
            for row in src_ws.iter_rows():
                for cell in row:
                    new_cell = sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                    # 复制样式
                    new_cell.font = cell.font.copy()
                    new_cell.border = cell.border.copy()
                    new_cell.fill = cell.fill.copy()
                    new_cell.number_format = cell.number_format
                    new_cell.protection = cell.protection.copy()
                    new_cell.alignment = cell.alignment.copy()

            # 复制列宽
            for col in src_ws.columns:
                column_letter = get_column_letter(col[0].column)  # 使用 get_column_letter 获取列字母
                sheet.column_dimensions[column_letter].width = src_ws.column_dimensions[column_letter].width

            # 复制行高
            for row in src_ws.iter_rows():
                sheet.row_dimensions[row[0].row].height = src_ws.row_dimensions[row[0].row].height

            # 复制合并单元格
            for merged_cell_range in src_ws.merged_cells.ranges:
                sheet.merge_cells(str(merged_cell_range))

        # 保存修改后的工作簿
        dest_wb.save(dest_file_path)
        print(f"成功将 {src_file_path} 的第 {sheet_i + 1} 个工作表复制到 {dest_file_path} 的所有工作表中，并保持格式不变")
    except Exception as ex:
        print(f"发生错误: {ex}")
    finally:
        # 关闭工作簿
        src_wb.close()
        dest_wb.close()


def modify_sheet_cell_value(file_path, cell='K1', sheet_id=[0], new_value=None, time_list=[], split_y="年", split_m="月", split_d="日"):
    """
    修改指定文件中的指定表格的内容
    :param file_path: 文件路径
    :param cell: 指定的表格
    :param sheet_id: 指定工作表的ID，若为[0]则在所有工作表的相应表格添加，若为列表[1,2,3]则在指定的工作表中添加
    :param new_value: 加入的内容
    :param time_list: 时间，初始年月日，结束年月日
    :param split_y: 年份后的间隔
    :param split_m: 月份后的间隔
    :param split_d: 日期后的间隔
    :return:
    """
    try:
        # 加载工作簿
        wb = load_workbook(filename=file_path)
        sheet_count = len(wb.worksheets)

        if new_value is not None and time_list==[]:
            if sheet_id==[0]:
                for sheet_i in range(sheet_count):
                    ws = wb.worksheets[sheet_i]
                    # 修改指定单元格的值
                    ws[cell] = new_value
                    print("写入: ", ws[cell])
            else:
                for sheet_i in sheet_id:
                    ws = wb.worksheets[sheet_i-1]
                    # 修改指定单元格的值
                    ws[cell] = new_value
                    print("写入: ", ws[cell])

        else:
            new_value = new_value or ""
            start_year, start_month, start_day, end_year, end_month, end_day = time_list

            if start_month == 0:
                if start_year > end_year:
                    print("开始年份不能晚于结束年份")
                    return

                current_year = start_year
                num = 0
                while current_year <= end_year:
                    title_name = str(current_year) + split_y
                    ws = wb.worksheets[num]
                    ws[cell] = new_value + title_name
                    print("写入: ", ws[cell])
                    current_year += 1
                    num += 1
                    if num > sheet_count-1:
                        break

            elif start_day == 0:
                # 确保开始日期在结束日期之前或相等
                if start_year > end_year or (start_year == end_year and start_month > end_month):
                    print("开始月份不能晚于结束月份")
                    return

                current_year, current_month = start_year, start_month
                num = 0
                while (current_year < end_year) or (current_year == end_year and current_month <= end_month):
                    title_name = str(current_year) + split_y + str(current_month) + split_m
                    ws = wb.worksheets[num]
                    ws[cell] = new_value + title_name
                    print("写入: ", ws[cell])
                    # 更新当前月份和年份
                    if current_month == 12:
                        current_month = 1
                        current_year += 1
                    else:
                        current_month += 1
                    num += 1
                    if num > sheet_count-1:
                        break

            else:
                if end_day == 0:
                    end_day = start_day
                # 将日期元素转换为日期对象
                start_date = datetime(start_year, start_month, start_day)
                end_date = datetime(end_year, end_month, end_day)

                # 确保开始日期在结束日期之前
                if start_date > end_date:
                    print("开始日期不能晚于结束日期")
                    return

                current_date = start_date
                num = 0
                while current_date <= end_date:
                    year, mouth, day = current_date.year, current_date.month, current_date.day
                    title_name = str(year) + split_y + str(mouth) + split_m + str(day) + split_d
                    ws = wb.worksheets[num]
                    ws[cell] = new_value + title_name
                    print("写入: ", ws[cell])
                    current_date += timedelta(days=1)
                    num += 1
                    if num > sheet_count-1:
                        break

        # 保存更改
        wb.save(file_path)
        print(f"成功修改 {file_path} 中第工作表的 {cell} 的值为: {new_value}")
    except Exception as ex:
        print(f"发生错误: {ex}")
    finally:
        # 关闭工作簿以释放资源
        wb.close()

def copy_and_rename_excel_files(file_path, num_copies=1, time_list=[], split_y="_", split_m="_", split_d=""):
    """
    复制指定数量的Excel文件，并默认复制一个文件，可以自定义文件的数量和是否以时间为后缀
    参数:
    file_path (str): 源Excel文件路径。
    num_copies (int): 要复制的文件数量，默认为1。
    """
    try:
        # 确保源文件存在
        if not os.path.isfile(file_path):
            print(f"源文件 {file_path} 不存在")
            return

        # 获取源文件所在的目录和文件名（不带扩展名）
        source_dir = os.path.dirname(file_path)
        source_filename, ext = os.path.splitext(os.path.basename(file_path))

        if time_list == []:
            for i in range(1, num_copies + 1):
                # 生成新的文件名
                new_filename = f"{source_filename}_{i}{ext}"
                # 构建新的文件路径
                new_file_path = os.path.join(source_dir, new_filename)

                # 复制文件并重命名
                shutil.copy(file_path, new_file_path)
                print(f"已复制并重命名为: {new_file_path}")

        else:
            start_year, start_month, start_day, end_year, end_month, end_day = time_list

            if start_month == 0:
                if start_year > end_year:
                    print("开始年份不能晚于结束年份")
                    return

                current_year = start_year
                while current_year <= end_year:
                    title_name = str(current_year) + split_y
                    new_filename = f"{source_filename}_{title_name}{ext}"
                    # 构建新的文件路径
                    new_file_path = os.path.join(source_dir, new_filename)
                    # 复制文件并重命名
                    shutil.copy(file_path, new_file_path)
                    print(f"已复制并重命名为: {new_file_path}")
                    current_year += 1

            elif start_day == 0:
                # 确保开始日期在结束日期之前或相等
                if start_year > end_year or (start_year == end_year and start_month > end_month):
                    print("开始月份不能晚于结束月份")
                    return

                current_year, current_month = start_year, start_month
                while (current_year < end_year) or (current_year == end_year and current_month <= end_month):
                    title_name = str(current_year) + split_y + str(current_month) + split_m
                    new_filename = f"{source_filename}_{title_name}{ext}"
                    # 构建新的文件路径
                    new_file_path = os.path.join(source_dir, new_filename)
                    # 复制文件并重命名
                    shutil.copy(file_path, new_file_path)
                    print(f"已复制并重命名为: {new_file_path}")
                    # 更新当前月份和年份
                    if current_month == 12:
                        current_month = 1
                        current_year += 1
                    else:
                        current_month += 1

            else:
                if end_day == 0:
                    end_day = start_day
                # 将日期元素转换为日期对象
                start_date = datetime(start_year, start_month, start_day)
                end_date = datetime(end_year, end_month, end_day)

                # 确保开始日期在结束日期之前
                if start_date > end_date:
                    print("开始日期不能晚于结束日期")
                    return

                current_date = start_date
                while current_date <= end_date:
                    year, mouth, day = current_date.year, current_date.month, current_date.day
                    title_name = str(year) + split_y + str(mouth) + split_m + str(day) + split_d
                    new_filename = f"{source_filename}_{title_name}{ext}"
                    # 构建新的文件路径
                    new_file_path = os.path.join(source_dir, new_filename)
                    # 复制文件并重命名
                    shutil.copy(file_path, new_file_path)
                    print(f"已复制并重命名为: {new_file_path}")
                    current_date += timedelta(days=1)

        print("所有文件复制完成")

    except Exception as ex:
        print(f"发生错误: {ex}")


def excel_cells_to_list(src_file_path, dest_file_path, cell_range, sheet=0, cell='E10'):
    """
    将Excel文件中指定工作表指定单元格范围的内容转换为列表，并将结果写入目标文件的指定单元格。

    参数:
    src_file_path (str): 源Excel文件路径。
    dest_file_path (str): 目标Excel文件路径。
    cell_range (str): 单元格范围，例如 'A1:A5'。
    sheet (int or str): 工作表的索引或名称，默认为第一个工作表。
    cell (str): 目标文件中要修改的单元格地址，例如 'E10'。
    """

    # 判断文件是否存在
    exist_src = os.path.exists(src_file_path)
    exist_dest = os.path.exists(dest_file_path)
    if exist_src and exist_dest:
        pass
    else:
        print("文件不存在")
        return

    # 加载源Excel文件
    workbook = openpyxl.load_workbook(src_file_path, data_only=True)

    # 选择工作表
    if isinstance(sheet, int):
        sheet = workbook.worksheets[sheet]  # 通过索引选择工作表
    else:
        sheet = workbook[sheet]  # 通过名称选择工作表

    # 初始化一个空列表来存储单元格内容
    cell_values = []

    # 解析单元格范围
    start_cell, end_cell = cell_range.split(':')

    # 提取起始和结束的列字母和行号
    start_col = re.findall(r'[A-Za-z]+', start_cell)[0]
    start_row = int(re.findall(r'\d+', start_cell)[0])
    end_col = re.findall(r'[A-Za-z]+', end_cell)[0]
    end_row = int(re.findall(r'\d+', end_cell)[0])

    # 遍历指定范围的单元格
    for row in range(start_row, end_row + 1):
        cell_address = f'{start_col}{row}'  # 构建单元格地址
        cell_value = sheet[cell_address].value  # 获取单元格的值
        cell_values.append(cell_value)

    print("提取的单元格值:", cell_values)

    # 将提取的值写入目标文件的指定单元格
    for i, value in enumerate(cell_values):
        if value is not None:  # 仅处理非空值
            print(f"正在写入值 '{value}' 到目标文件的单元格 {cell}")
            modify_sheet_cell_value(
                file_path=dest_file_path,
                cell=cell,
                sheet_id=[i + 1],  # 假设 sheet_id 是目标工作表的索引
                new_value=str(value)
            )

    # 保存并关闭工作簿
    workbook.close()


def format_excel_cell(workbook, cell_address, sheet_i, alignment=None, width=None, height=None):
    """
    更改Excel文件中指定单元格的格式。
    参数:
    file_path (str): Excel文件路径。
    cell_address (str): 单元格地址，例如 'A1'。
    sheet_i: 工作表序号
    alignment (str): 对齐方式，可选 'left', 'center', 'right'。
    width (float): 列宽。
    height (float): 行高。
    """
    # 加载Excel文件
    #workbook = openpyxl.load_workbook(file_path)

    # 选择工作表
    sheet = workbook.worksheets[sheet_i]

    # 获取指定单元格
    cell = sheet[cell_address]

    # 设置对齐方式
    if alignment:
        if alignment == 'left':
            cell.alignment = Alignment(horizontal='left')
        elif alignment == 'center':
            cell.alignment = Alignment(horizontal='center')
        elif alignment == 'right':
            cell.alignment = Alignment(horizontal='right')

    # 设置列宽
    if width:
        column_letter = get_column_letter(cell.column)
        sheet.column_dimensions[column_letter].width = width

    # 设置行高
    if height:
        sheet.row_dimensions[cell.row].height = height

    print(f"单元格 {cell_address} 的格式已更新。")


def format_excel_cell_range(file_path, cell_address, sheet_id=[0], alignment=None, width=None, height=None):
    """
    更改Excel文件中指定工作表的指定单元格的格式。
    参数:
    file_path (str): Excel文件路径。
    cell_address (str): 单元格地址，例如 'A1'。
    sheet_id: 指定工作表的ID，若为[0]则在所有工作表的相应表格添加，若为列表[1,2,3]则在指定的工作表中添加
    alignment (str): 对齐方式，可选 'left', 'center', 'right'。
    width (float): 列宽。
    height (float): 行高。
    """
    # 加载Excel文件
    workbook = openpyxl.load_workbook(file_path)

    # 计算工作表数量
    sheet_count = len(workbook.worksheets)
    print(sheet_count)

    if sheet_id == [0]:
        for sheet_i in range(sheet_count):
            format_excel_cell(
                workbook=workbook,
                sheet_i=sheet_i,
                cell_address=cell_address,
                alignment=alignment,
                width=width,
                height=height
            )
    else:
        for sheet_i in sheet_id:
            if sheet_i > sheet_count:
                print(f"工作表 {sheet_i} 不存在。")
                continue
            format_excel_cell(
                workbook=workbook,
                sheet_i=sheet_i-1,
                cell_address=cell_address,
                alignment=alignment,
                width=width,
                height=height
            )

    # 保存文件
    workbook.save(file_path)
    print(f"Excel文件 {file_path} 已保存。")
    # 保存并关闭工作簿
    workbook.close()


def merge_excel_cells(file_path, start_cell, end_cell, sheet_id=[0]):
    """
    合并 Excel 文件中的指定单元格
    :param file_path: Excel 文件路径
    :param start_cell: 合并区域的起始单元格（例如 'A1'）
    :param end_cell: 合并区域的结束单元格（例如 'B2'）
    :param sheet_id: 指定工作表的ID，若为[0]则在所有工作表的相应表格添加，若为列表[1,2,3]则在指定的工作表中添加
    """
    # 加载 Excel 文件
    workbook = openpyxl.load_workbook(file_path)

    # 计算工作表数量
    sheet_count = len(workbook.worksheets)
    print(sheet_count)

    if sheet_id == [0]:
        for sheet_i in range(sheet_count):
            # 选择工作表
            sheet = workbook.worksheets[sheet_i]
            # 合并单元格
            sheet.merge_cells(start_cell + ':' + end_cell)
    else:
        for sheet_i in sheet_id:
            if sheet_i > sheet_count:
                print(f"工作表 {sheet_i} 不存在。")
                continue
            # 选择工作表
            sheet = workbook.worksheets[sheet_i-1]
            # 合并单元格
            sheet.merge_cells(start_cell + ':' + end_cell)

    # 保存修改后的 Excel 文件
    workbook.save(file_path)
    print(f"单元格 {start_cell} 到 {end_cell} 已成功合并！")


def process_path_or_filename(input_str):
    """
    如果输入的是一个文件名，则直接返回。
    如果输入的是Windows路径，则转换为Python友好的路径格式。
    然后对结果执行下一步操作。
    """
    # 检查是否是简单的文件名（不包含任何路径分隔符）
    if os.sep not in input_str and '/' not in input_str and '\\' not in input_str:
        print("Detected simple filename.")
    else:
        # 转换Windows风格路径到Python友好的路径格式
        input_str = input_str.replace('\\', '/')
        print("Converted to Python-friendly path format.")
    return input_str

def write_df_to_excel(output_filename, df):
    """
    将给定的DataFrame写入到指定的Excel文件中。
    参数:
    df (pd.DataFrame): 要写入Excel文件的pandas DataFrame。
    output_filename (str): 输出的Excel文件名，包括路径。
    """
    try:
        df.to_excel(output_filename, index=False)
        print(f"文件成功写入到文件: {output_filename}")
    except Exception as e:
        print(f"写入文件时出错: {e}")


def process_excel_files(file_list, keyword=None, cell_position=None, operation='read', direction=None, num_of_cells=1,
                        output_file='output_m.xlsx', extract_formula=False, transpose=False):
    """
        处理给定的Excel文件列表，查找包含特定关键字或指定位置的单元格，并对其周围的单元格或自身进行操作。
        :param file_list: 包含Excel文件路径的列表
        :param keyword: 查找的关键字（可选）
        :param cell_position: 单元格位置，如'A1'（可选）
        :param operation: 'read' 表示读取周围单元格，'modify' 表示修改当前单元格
        :param direction: 要提取的内容方向，可以是'up', 'down', 'left', 'right'
        :param num_of_cells: 提取多少个相邻单元格（仅当operation为'read'时有效）
        :param output_file: 输出结果的Excel文件名
        :param extract_formula: 是否提取单元格中的公式，默认为False
        :param transpose: 是否要将提取到的数据进行转置，默认为False
    """
    all_results = []
    all_data = []

    for file in file_list:
        wb = load_workbook(filename=file, data_only=not extract_formula)  # 如果不提取公式，则data_only=True

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            if cell_position and not direction:  # 当提供了单元格位置但未提供方向时
                col, row = column_index_from_string(cell_position[0]), int(cell_position[1:])
                cell = ws.cell(row=row, column=col)

                result = {'File': file, 'Sheet': sheet_name, 'Keyword Position': cell_position}
                extracted_data = [cell.value if not extract_formula else cell.formula]
                all_data.append(extracted_data)
                all_results.append(result)

            elif cell_position and direction:  # 当提供了单元格位置和方向时
                col, row = column_index_from_string(cell_position[0]), int(cell_position[1:])
                cell = ws.cell(row=row, column=col)

                result = {'File': file, 'Sheet': sheet_name, 'Keyword Position': cell_position}
                extracted_data = []
                if direction == 'right':
                    for i in range(1, num_of_cells + 1):
                        target_cell = ws.cell(row=row, column=col + i)
                        if target_cell:
                            extracted_data.append(target_cell.value if not extract_formula else target_cell.formula)
                elif direction == 'left':
                    for i in range(num_of_cells):
                        target_cell = ws.cell(row=row, column=col - i - 1)
                        if target_cell:
                            extracted_data.append(target_cell.value if not extract_formula else target_cell.formula)
                elif direction == 'up':
                    for i in range(num_of_cells):
                        target_cell = ws.cell(row=row - i - 1, column=col)
                        if target_cell:
                            extracted_data.append(target_cell.value if not extract_formula else target_cell.formula)
                elif direction == 'down':
                    for i in range(1, num_of_cells + 1):
                        target_cell = ws.cell(row=row + i, column=col)
                        if target_cell:
                            extracted_data.append(target_cell.value if not extract_formula else target_cell.formula)
                all_data.append(extracted_data)
                all_results.append(result)

            elif keyword:
                for row in ws.iter_rows():
                    for cell in row:
                        if str(cell.value) == keyword:
                            result = {'File': file, 'Sheet': sheet_name,
                                      'Keyword Position': f'{cell.row}, {cell.column}'}

                            if operation == 'read':
                                extracted_data = []
                                if direction == 'right':
                                    for i in range(1, num_of_cells + 1):
                                        target_cell = ws.cell(row=cell.row, column=cell.column + i)
                                        if target_cell:
                                            extracted_data.append(
                                                target_cell.value if not extract_formula else target_cell.formula)
                                elif direction == 'left':
                                    for i in range(num_of_cells):
                                        target_cell = ws.cell(row=cell.row, column=cell.column - i - 1)
                                        if target_cell:
                                            extracted_data.append(
                                                target_cell.value if not extract_formula else target_cell.formula)
                                elif direction == 'up':
                                    for i in range(num_of_cells):
                                        target_cell = ws.cell(row=cell.row - i - 1, column=cell.column)
                                        if target_cell:
                                            extracted_data.append(
                                                target_cell.value if not extract_formula else target_cell.formula)
                                elif direction == 'down':
                                    for i in range(1, num_of_cells + 1):
                                        target_cell = ws.cell(row=cell.row + i, column=cell.column)
                                        if target_cell:
                                            extracted_data.append(
                                                target_cell.value if not extract_formula else target_cell.formula)
                                all_data.append(extracted_data)

                            elif operation == 'modify':
                                cell.value = "Modified"
                                result['Modification'] = "Cell value modified."

                            all_results.append(result)

            if operation == 'modify':
                wb.save(file)

    # 将所有结果保存到新的Excel文件中
    df_results = pd.DataFrame(all_results)
    df_data = pd.DataFrame(all_data)
    print(df_data)
    # 将列的索引改为从 1 开始
    df_data.columns = range(1, len(df_data.columns) + 1)

    # 将两个DataFrame合并
    df_results = pd.concat([df_results, df_data], axis=1)

    if transpose:
        df_results = df_results.T
        df_data = df_data.T

    print(df_results)
    print(df_data)

    # 尝试将 DataFrame 的内容转换为 double 类型
    try:
        # 使用 astype() 将 DataFrame 转换为 float64 类型
        df_data = df_data.astype(float)
    except ValueError as e:
        # 如果转换失败，捕获异常并提示
        print(f"转换失败: {e}")
        print("请检查数据中是否包含非数值型数据。")

    # 计算每一列的和
    sum_row = df_data.sum(numeric_only=True).rename('总和')

    df_data_describe = df_data.describe()
    # 将求和结果添加到描述性统计的最后一行
    df_data_describe = df_data_describe._append(sum_row)
    print(df_data_describe)

    time_1 = get_current_date()

    # 判断output_file是否存在
    if os.path.exists(output_file):
        print("文件已存在，将覆盖")
        made = 'a'
    else:
        print("文件不存在，将创建")
        made = 'w'

    with pd.ExcelWriter(output_file, engine='openpyxl', mode=made) as writer:
        df_results.to_excel(writer, index=False, sheet_name='相关数据' + str(time_1))
        df_data_describe.to_excel(writer, index=True, sheet_name='汇总结果' + str(time_1))


if __name__ == '__main__':
    '''
    current_time = get_current_date()

    src_file_path = '运营情况1.xlsx'  # 源Excel文件路径
    dest_file_path = '入库单.xlsx'  # 目标Excel文件路径
    cell_range = 'T5:T35'  # 单元格范围

    copy_first_sheet_to_all_sheets(src_file_path='表1.xlsx', dest_file_path='入库单.xlsx')
    '''
    copy_first_sheet_to_all_sheets(src_file_path='入库单\\入库单_2024年10月.xlsx', dest_file_path='入库单\\tt1.xlsx', sheet_i=2)